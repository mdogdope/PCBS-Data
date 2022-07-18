import sqlite3
from openpyxl import load_workbook
import os

def replaceSpaces(text):
	ret = ""
	for ch in text:
		if(ch == " "):
			ret += "_"
		else:
			ret += ch
	return ret

def getType(text):
	if type(text) == int:
		return "int"
	else:
		return "varchar"

wb = load_workbook(filename="PCBS_Parts.xlsx")

if os.path.exists("pcbs.db"):
	os.remove("pcbs.db")

con = sqlite3.connect("pcbs.db")
cur = con.cursor()

for sheetName in wb.sheetnames:
	sheet = wb[sheetName]
	maxRow = sheet.max_row + 1
	maxCol = sheet.max_column + 1
	
	# if sheetName == "Cases":
	# 	continue
	
	# Create table for sheet
	baseCreate = "CREATE TABLE " + replaceSpaces(sheetName) + " ({});"
	cols = ""
	types = []
	for col in range(1, maxCol):
		if col > 1:
			cols += ","
		cols += '"' + replaceSpaces(sheet.cell(1,col).value) + '" ' + getType(sheet.cell(2,col).value)
		types.append(getType(sheet.cell(2, col).value))
		pass
	
	cmd = baseCreate.format(cols)
	
	print(cmd)
	cur.execute(cmd)
	
	
	
	# Populate data
	baseInsert = "INSERT INTO " + replaceSpaces(sheetName) + " VALUES ({});"
	for row in range(2, maxRow):
		data = ""
		for col in range(1,maxCol):
			if col > 1:
				data += ","
			if(types[col - 1] == "varchar"):
				data += '"'
			
			data += str(sheet.cell(row,col).value)
			
			if(types[col - 1] == "varchar"):
				data += '"'
			pass
		
		cmd = baseInsert.format(data)
		
		if sheetName == "Motherboards":
			# print(cmd)
			pass
		cur.execute(cmd)
		pass
	
	con.commit()
	pass

